from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os

directory_path = 'data/'
path_result = 'result_ticket.xlsx'

"""
top読み込み
"""
sheet_name = 'top'
excel_files = [f for f in os.listdir(directory_path) if f.endswith('.xlsx')]
df = pd.DataFrame(columns=[])
for file in excel_files:
    file_path = os.path.join(directory_path, file)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    new_row = {  'チケットNo':os.path.splitext(file)[0]\
		,'状況':sheet['B3'].value\
		,'期限':sheet['B4'].value\
		,'担当':sheet['B6'].value\
               }
    new_row_df = pd.DataFrame([new_row])
    df = pd.concat([df, new_row_df], ignore_index=True)
#print(df)

"""
出力
"""
#df.to_excel('out.xlsx', index=False, engine='openpyxl')
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
ws.auto_filter.ref = ws.dimensions
wb.save(path_result)
